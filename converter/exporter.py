import shutil
import os
import openpyxl
import logging
from copy import copy
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger("converter.exporter")

WARNING_KEYWORDS = ("inter-state transaction", "gstin lookup failed")

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

SUGGESTED_FIXES = {
    "gstin of supplier is blank": "Fill in the supplier GSTIN from the source invoice.",
    "invalid gstin format": "Correct the GSTIN to match the 15-character format.",
    "invoice number is blank": "Enter the invoice number as printed on the source document.",
    "invoice date is blank": "Enter the invoice date in DD-MM-YYYY format.",
    "invalid invoice date format": "Re-enter the date as DD-MM-YYYY.",
    "gst rate must be numeric": "Correct the GST rate cell to a plain number.",
    "is not supported by tally template": "Verify the rate against the source invoice; only 0/5/12/18/28% are supported.",
    "taxable value must be numeric": "Correct the taxable value cell to a plain number.",
    "taxable value is negative": "Confirm whether this is a valid credit note; otherwise correct the value.",
    "duplicate invoice detected": "Check the source file for a duplicated row and remove it if erroneous.",
    "does not match expected": "Recheck CGST/SGST against the taxable value and rate for calculation errors.",
    "inter-state transaction": "Tally template has no IGST columns; this row was skipped from Sheet2.",
    "gstin lookup failed": "Party name could not be auto-resolved; row was still imported with a fallback name. Verify the GSTIN on GSTzen and correct the party name in Sheet1.",
}

def classify_error(error_text):
    """Returns (severity, fill) for a given error description string."""
    text = error_text.lower()
    if any(kw in text for kw in WARNING_KEYWORDS):
        return "Warning", YELLOW_FILL
    return "Critical", RED_FILL

def suggest_fix(error_text):
    """Looks up a human-readable suggested fix based on keywords in the error text."""
    text = error_text.lower()
    for keyword, fix in SUGGESTED_FIXES.items():
        if keyword in text:
            return fix
    return "Review the row manually against the source GSTR data."

def export_to_excel(mapped_rows, errors, template_path, output_path):
    """
    Copies the Tally import template to output path, overwrites Sheet2 with converted transactions,
    and adds an 'Errors' sheet listing any records that failed validation.
    Preserves all template formatting, fonts, and properties.
    """
    # Create output directory if it doesn't exist
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        
    logger.info(f"Copying template from {template_path} to {output_path}...")
    shutil.copy(template_path, output_path)
    
    # Load workbook using openpyxl
    wb = openpyxl.load_workbook(output_path)
    
    if "Sheet2" not in wb.sheetnames:
        raise ValueError("Target worksheet 'Sheet2' not found in Tally template.")
        
    ws = wb["Sheet2"]

    # Add "Purchase IGST @ Rate" / "Input IGST @ Rate" header columns per rate slab.
    # These reuse blank template columns (22-29), styled to match the existing
    # Purchase/Input columns (e.g. "Purchase GST @ 5%" / "Input Cgst @ 2.5%") exactly.
    purchase_header_ref = ws.cell(row=1, column=9)   # "Purchase GST @ 5%"
    input_header_ref = ws.cell(row=1, column=10)     # "Input Cgst @ 2.5%"
    purchase_data_ref = ws.cell(row=2, column=9)
    input_data_ref = ws.cell(row=2, column=10)
    ref_width = ws.column_dimensions[openpyxl.utils.get_column_letter(9)].width

    for rate, (p_col, i_col) in {5: (22, 23), 12: (24, 25), 18: (26, 27), 28: (28, 29)}.items():
        p_header = ws.cell(row=1, column=p_col, value=f"Purchase IGST @ {rate}%")
        p_header.font = copy(purchase_header_ref.font)
        p_header.fill = copy(purchase_header_ref.fill)
        p_header.border = copy(purchase_header_ref.border)
        p_header.alignment = copy(purchase_header_ref.alignment)

        i_header = ws.cell(row=1, column=i_col, value=f"Input IGST @ {rate}%")
        i_header.font = copy(input_header_ref.font)
        i_header.fill = copy(input_header_ref.fill)
        i_header.border = copy(input_header_ref.border)
        i_header.alignment = copy(input_header_ref.alignment)

        ws.column_dimensions[openpyxl.utils.get_column_letter(p_col)].width = ref_width
        ws.column_dimensions[openpyxl.utils.get_column_letter(i_col)].width = ref_width

    # Cache the styles of row 2 (the first data row) to apply to all newly added rows
    logger.info("Caching styling from first data row...")
    style_cache = {}
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col_idx)
        style_cache[col_idx] = {
            "font": copy(cell.font) if cell.font else None,
            "fill": copy(cell.fill) if cell.fill else None,
            "border": copy(cell.border) if cell.border else None,
            "alignment": copy(cell.alignment) if cell.alignment else None,
            "number_format": cell.number_format
        }

    # Override the new IGST data-column styles to match their reference columns
    # (the blank template columns don't carry the accounting number format).
    for p_col, i_col in [(22, 23), (24, 25), (26, 27), (28, 29)]:
        style_cache[p_col] = {
            "font": copy(purchase_data_ref.font),
            "fill": copy(purchase_data_ref.fill),
            "border": copy(purchase_data_ref.border),
            "alignment": copy(purchase_data_ref.alignment),
            "number_format": purchase_data_ref.number_format
        }
        style_cache[i_col] = {
            "font": copy(input_data_ref.font),
            "fill": copy(input_data_ref.fill),
            "border": copy(input_data_ref.border),
            "alignment": copy(input_data_ref.alignment),
            "number_format": input_data_ref.number_format
        }

    # Clear existing data rows (from row 2 onwards)
    max_row = ws.max_row
    if max_row >= 2:
        logger.info(f"Clearing {max_row - 1} sample data rows from Sheet2...")
        ws.delete_rows(2, max_row - 1)
        
    # Write mapped transaction rows and apply styling
    logger.info(f"Writing {len(mapped_rows)} mapped rows to Sheet2...")
    for r_idx, row_data in enumerate(mapped_rows, start=2):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            
            # Apply cached cell styling
            if c_idx in style_cache:
                styles = style_cache[c_idx]
                if styles["font"]:
                    cell.font = styles["font"]
                if styles["fill"]:
                    cell.fill = styles["fill"]
                if styles["border"]:
                    cell.border = styles["border"]
                if styles["alignment"]:
                    cell.alignment = styles["alignment"]
                if styles["number_format"]:
                    cell.number_format = styles["number_format"]
                    
    # Setup the Errors sheet - only present in the final workbook if errors actually exist.
    if errors:
        logger.info("Setting up the Errors worksheet...")
        if "Errors" in wb.sheetnames:
            ws_err = wb["Errors"]
            ws_err.delete_rows(1, ws_err.max_row)
        else:
            ws_err = wb.create_sheet("Errors")

        # Define fonts
        bold_font = Font(name="Aptos Narrow", size=11, bold=True)
        normal_font = Font(name="Aptos Narrow", size=11)

        # Write headers for Errors worksheet - full record context so every failed
        # invoice is traceable back to the source row without reopening the input file.
        headers = ["Row Number", "Invoice Number", "GSTIN", "Party Name", "Invoice Date",
                   "GST Rate", "Taxable Value", "Issue", "Severity", "Suggested Fix"]
        ws_err.append(headers)
        for col_idx in range(1, len(headers) + 1):
            ws_err.cell(row=1, column=col_idx).font = bold_font

        logger.info(f"Writing {len(errors)} error rows to Errors sheet...")
        for r_idx, err in enumerate(errors, start=2):
            severity, fill = classify_error(err["error"])
            fix = suggest_fix(err["error"])

            ws_err.cell(row=r_idx, column=1, value=err["row"]).font = normal_font
            ws_err.cell(row=r_idx, column=2, value=err["invoice"]).font = normal_font
            ws_err.cell(row=r_idx, column=3, value=err["gstin"]).font = normal_font
            ws_err.cell(row=r_idx, column=4, value=err.get("party_name") or "Party Name Not Found").font = normal_font
            ws_err.cell(row=r_idx, column=5, value=err.get("invoice_date") or "N/A").font = normal_font
            ws_err.cell(row=r_idx, column=6, value=err.get("rate") if err.get("rate") is not None else "N/A").font = normal_font
            ws_err.cell(row=r_idx, column=7, value=err.get("taxable_value") if err.get("taxable_value") is not None else "N/A").font = normal_font
            ws_err.cell(row=r_idx, column=8, value=err["error"]).font = normal_font
            severity_cell = ws_err.cell(row=r_idx, column=9, value=severity)
            severity_cell.font = normal_font
            severity_cell.fill = fill
            ws_err.cell(row=r_idx, column=10, value=fix).font = normal_font

        # Automatically set column widths for Errors sheet
        for col in ws_err.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_err.column_dimensions[col_letter].width = max(max_len + 3, 12)
    else:
        logger.info("No errors encountered. Omitting the Errors worksheet.")
        if "Errors" in wb.sheetnames:
            del wb["Errors"]

    # Final workbook structure: drop the legacy sample-data "Sheet1" that ships with the
    # template, then rename the populated "Sheet2" to "Sheet1" so the output only ever
    # contains "Sheet1" (the converted data) and, when applicable, "Errors".
    if "Sheet1" in wb.sheetnames and wb["Sheet1"] is not ws:
        del wb["Sheet1"]
    ws.title = "Sheet1"

    # Save the workbook
    wb.save(output_path)
    logger.info(f"Output workbook successfully saved to {output_path}")

if __name__ == "__main__":
    # Test exporter
    logging.basicConfig(level=logging.INFO)
    test_rows = [[f"GSTIN_{i}", f"INV_{i}", "01-05-2025", "Test Party", "NARRATION", "Purchase New", 18, 0,0,0,0,0,0,0, 100, 9, 9, 0,0,0,0] + [None]*59 + ["Status Done"] for i in range(5)]
    test_errors = [{"row": 15, "invoice": "INV_ERR_1", "gstin": "27ERR", "error": "Invalid Rate"}]
    export_to_excel(test_rows, test_errors, "templates/PURCHASE IMPORT MAY.xlsx", "output/test_output.xlsx")
